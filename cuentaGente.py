#Cuenta los hombres y mujeres separados por edad
import os
import openpyxl
import numpy as np
os.chdir('Aqui tu ruta')#Todos Juntos en la misma carpeta, el Script y los archivos de Excel. 

def Edad_Y_Sexo(fileName):
    workbook = openpyxl.load_workbook(fileName+'.xlsx')
    sheet1 = workbook.get_active_sheet()
    #variables
    nfemtotal=0 
    nfem18a29 = 0 
    nfem30a44 = 0
    nfem45a59 = 0
    nfem60ymas = 0
    sindato = 0
    nmasc_total = 0
    nmasc_18a29 = 0
    nmasc_30a44 = 0
    nmasc_45a59 = 0
    nmasc_60ymas = 0
    for i in range(2,sheet1.max_row + 1):
        sexo = sheet1.cell(row = i, column = 7).value
        edad = sheet1.cell(row = i, column = 8).value
    #Que tal si no hay datos? Los hacemos Cero
        if sexo == None:
            sexo = 'cero'
        if edad == None:
            edad = 0
    #CUENTA MUJERES
        # toma todas las mujeres
        if  sexo == 'FEMENINO' or sexo == 'F' or sexo == 'f' or sexo == 'Femenino' or sexo == 'femenino' or sexo == 'MUJER' or sexo == 'Mujer' or sexo == 'mujer':
                nfemtotal= nfemtotal + 1
                #De cada Mujer, revisa su edad
                if  int(edad) > 17 and int(edad) < 30:
                    nfem18a29 = nfem18a29 + 1
                elif int(edad) > 29 and int(edad) < 45:
                    nfem30a44 = nfem30a44 + 1
                elif int(edad) > 44 and int(edad) < 60:
                    nfem45a59 = nfem45a59 + 1
                elif int(edad) > 59:
                    nfem60ymas = nfem60ymas + 1
                else:
                    sindato = sindato + 1
    #CUENTA HOMBRES
        # toma todas los hombres
        if  sexo == 'MASCULINO' or sexo == 'M' or sexo == 'm' or sexo == 'Masculino' or sexo == 'masculino' or sexo == 'HOMBRE' or sexo == 'Hombre' or sexo == 'hombre' or sexo == 'H' or sexo == 'h': 
                nmasc_total= nmasc_total + 1
                #De cada Hombre, revisa su edad
                if int(edad) > 17 and int(edad) < 30:
                    nmasc_18a29 = nmasc_18a29 + 1
                elif int(edad) > 29 and int(edad) < 45:
                    nmasc_30a44 = nmasc_30a44 + 1
                elif int(edad) > 44 and int(edad) < 60:
                    nmasc_45a59 = nmasc_45a59 + 1
                elif int(edad) > 59:
                    nmasc_60ymas = nmasc_60ymas + 1
                else:
                    sindato = sindato + 1
                        

    #Output
    totalSumaF= nfem18a29 + nfem30a44 + nfem45a59 + nfem60ymas + sindato
    totalSumaM= nmasc_18a29 + nmasc_30a44 + nmasc_45a59 + nmasc_60ymas + sindato
    total= totalSumaF + totalSumaM
    print('-----------TOTALES------------')
    print('Hombres totales ',totalSumaM)
    print('Mujeres totales ',totalSumaF)
    print('-----------EDADES------------')
    print('Mujeres: ')
    print('de 18 a 29: ',nfem18a29)
    print('de 29 a 44: ',nfem30a44)
    print('de 45 a 59: ',nfem45a59)
    print('de 60 y mas: ',nfem60ymas)
    print('Hombres: ')
    print('de 18 a 29: ',nmasc_18a29)
    print('de 29 a 44: ',nmasc_30a44)
    print('de 45 a 59: ',nmasc_45a59)
    print('de 60 y mas: ',nmasc_60ymas)
    print('Sin Datos: ',sindato)
    print('max row ',sheet1.max_row)
    print('Total ', total)
    print('----------------------------')
    print('Guardar datos en Tabla de excel? responde: si/no')
    resp = input()
    if resp == 'si' or resp == 's' or resp == 'Si'or resp == 'S':
        #guardamos datos en tabla
        wb = openpyxl.Workbook()
        #newsheet = wb.get_sheet_by_name('Sheet')
        newsheet= wb['Sheet']
        newsheet['A1'] = 'EDAD'
        newsheet['A2'] = '18 a 29'
        newsheet['A3'] = '29 a 44'
        newsheet['A4'] = '45 a 59'
        newsheet['A5'] = '60 y mas'
        newsheet['A6'] = 'Sin dato'

        newsheet['B1'] = 'HOMBRES'
        newsheet['B2'] = nmasc_18a29
        newsheet['B3'] = nmasc_30a44
        newsheet['B4'] = nmasc_45a59
        newsheet['B5'] = nmasc_60ymas

        newsheet['C1'] = 'MUJERES'
        newsheet['C2'] = nfem18a29
        newsheet['C3'] = nfem30a44
        newsheet['C4'] = nfem45a59
        newsheet['C5'] = nfem60ymas
        newsheet['C6'] = sindato
        print('Escribe el nombre que deseas para guardar el archivo:')
        nombre = input()
        wb.save(nombre+'.xlsx')
        print('guardado en como '+ nombre+'.xlsx')
    else:
        print('Adios')

    return
print('Instrucciones: Este programa te pedirá los archivos a leer de uno por uno, cada que acabe uno te preguntará si deseas guardarlo')    

print('Cuantos archivos son?')
archivos = int(input())

for x in range(1,archivos+1):
    print('escribe el nombre del archivo numero', x)
    filename = input()
    Edad_Y_Sexo(filename)

    


    
